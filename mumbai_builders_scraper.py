"""
Mumbai Builders Scraper — FINAL (exact selectors from real HTML)
=================================================================
HTML structure confirmed from proptiger.com:

  <div class="builder-projects-row">          ← one builder group
    <div class="project-mini-card">           ← one project
      <h4 class="project-name">Purva Panorama
      <span class="loc-text">Thane West, Mumbai
      <span class="price">1.4 Cr Onwards
      <span class="bhk">2,3 BHK
    <div class="builder-row-footer">
      <a class="builder-cta">View 7 Projects  ← builder name from href

NOTE: Area and Possession are NOT present on this listing page.
      proptiger only shows them on individual project detail pages.
      This scraper fetches each project's detail page to get them.

pip install requests beautifulsoup4 openpyxl
python mumbai_builders_scraper.py
"""

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import time
import re

# ----------------------------
TARGET_URL = "https://www.proptiger.com/mumbai/all-builders"
BASE_DOMAIN = "https://www.proptiger.com"
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
    "Accept-Encoding": "gzip, deflate, br",
    "Connection": "keep-alive",
    "Referer": "https://www.google.com/",
}
FETCH_DETAIL = True   # set False to skip detail pages (faster, but Area/Possession empty)
MAX_PAGES    = 726     # number of listing pages to scrape (each has ~200 projects)
# ----------------------------

session = requests.Session()


def fetch(url, retries=3):
    for attempt in range(retries):
        try:
            r = session.get(url, headers=HEADERS, timeout=15)
            if r.status_code == 200:
                return BeautifulSoup(r.text, "html.parser")
            print(f"  HTTP {r.status_code} — {url}")
        except Exception as e:
            print(f"  Attempt {attempt+1} failed: {e}")
        time.sleep(2)
    return None


def extract_text(tag):
    return tag.get_text(strip=True) if tag else ""


def get_builder_name_from_href(href):
    """
    Builder name is encoded in the href of .builder-cta link.
    e.g. /Mumbai/puravankara-limited-100046  →  Puravankara Limited
    """
    if not href:
        return ""
    # Get the slug part: puravankara-limited-100046
    slug = href.rstrip("/").split("/")[-1]
    # Remove trailing numeric ID: puravankara-limited
    slug = re.sub(r"-\d+$", "", slug)
    # Convert hyphens to spaces and title-case
    name = slug.replace("-", " ").title()
    return name


def fetch_project_detail(href):
    """
    Fetch individual project page to get Area and Possession.
    These fields only exist on detail pages, not the listing page.
    """
    url = BASE_DOMAIN + href if href.startswith("/") else href
    soup = fetch(url)
    if not soup:
        return "", ""

    area       = ""
    possession = ""

    # ── Area ─────────────────────────────────────────────────
    # proptiger detail page uses these selectors for area
    for sel in [
        "[class*='area']", "[class*='size']",
        "[class*='carpet']", "[class*='super-area']",
        "[class*='builtup']", "[class*='unit-size']",
    ]:
        tag = soup.select_one(sel)
        if tag:
            v = extract_text(tag)
            if re.search(r"\d", v) and re.search(r"sq|ft|m²", v, re.I):
                area = v
                break

    # Regex fallback across whole page
    if not area:
        page_text = soup.get_text(" ", strip=True)
        m = re.search(
            r"(\d[\d,\.]*\s*(?:[-–to]+\s*\d[\d,\.]*)?\s*"
            r"(?:sq\.?\s*ft\.?|sqft|sq\.?\s*m|sqm|m²))",
            page_text, re.I
        )
        if m:
            area = m.group(1).strip()

    # ── Possession ───────────────────────────────────────────
    for sel in [
        "[class*='possession']", "[class*='handover']",
        "[class*='status']",     "[class*='completion']",
        "[class*='occupancy']",
    ]:
        tag = soup.select_one(sel)
        if tag:
            v = extract_text(tag)
            if v:
                possession = v
                break

    if not possession:
        page_text = soup.get_text(" ", strip=True) if not area else soup.get_text(" ", strip=True)
        m = re.search(
            r"(Ready\s+to\s+Move(?:\s*[-–]\s*\w+)?"
            r"|Under\s+Construction"
            r"|New\s+Launch"
            r"|Possession\s+(?:by\s+)?[\w\s,\.]*?20\d\d"
            r"|(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)"
            r"[a-z]*[\s\'\-,]*20\d\d"
            r"|Q[1-4][\s\-\']*20\d\d)",
            soup.get_text(" ", strip=True), re.I
        )
        if m:
            possession = m.group(0).strip()

    return area, possession


def scrape_builders(base_url):
    records = []

    # Hit homepage first to get cookies
    session.get("https://www.proptiger.com", headers=HEADERS, timeout=15)
    time.sleep(1)

    for page in range(1, MAX_PAGES + 1):
        url = f"{base_url}?page={page}"
        print(f"\nFetching page {page} -> {url}")

        soup = fetch(url)
        if not soup:
            print("  Failed to fetch. Stopping.")
            break

        # Each builder group is a div.builder-projects-row
        builder_rows = soup.select("div.builder-projects-row")
        if not builder_rows:
            print("  No builder rows found. Stopping.")
            break

        print(f"  Found {len(builder_rows)} builder groups")

        for row in builder_rows:

            # ── Builder name from footer CTA link ─────────────
            cta = row.select_one("a.builder-cta")
            builder_href  = cta["href"] if cta and cta.get("href") else ""
            builder_name  = get_builder_name_from_href(builder_href)

            # ── Each project card inside this builder row ──────
            cards = row.select("div.project-mini-card")
            print(f"    Builder: {builder_name or '?'}  |  Projects: {len(cards)}")

            for card in cards:
                record = {
                    "Builder":    builder_name,
                    "Project":    "",
                    "Location":   "",
                    "Price":      "",
                    "BHK":        "",
                    "Area":       "",
                    "Possession": ""
                }

                # Project name — h4.project-name
                record["Project"] = extract_text(card.select_one("h4.project-name"))

                # Location — span.loc-text
                record["Location"] = extract_text(card.select_one("span.loc-text"))

                # Price — span.price
                record["Price"] = extract_text(card.select_one("span.price"))

                # BHK — span.bhk
                record["BHK"] = extract_text(card.select_one("span.bhk"))

                # ── Area & Possession from detail page ─────────
                if FETCH_DETAIL:
                    link_tag = card.select_one("a.no-ajaxy")
                    if link_tag and link_tag.get("href"):
                        project_href = link_tag["href"]
                        area, possession = fetch_project_detail(project_href)
                        record["Area"]       = area
                        record["Possession"] = possession
                        time.sleep(0.5)   # polite delay per detail page

                records.append(record)

        time.sleep(1)   # polite delay between listing pages

    return records


def save_to_excel(records):
    wb = Workbook()
    ws = wb.active
    ws.title = "Mumbai Builders"

    headers = ["Builder", "Project", "Location", "Price", "BHK", "Area", "Possession"]
    widths  = [25, 30, 25, 18, 10, 15, 20]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", start_color="1F4E79")
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin")
    )

    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.border    = border
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(i)].width = widths[i - 1]

    for r, rec in enumerate(records, 2):
        for c, h in enumerate(headers, 1):
            cell           = ws.cell(row=r, column=c, value=rec[h])
            cell.border    = border
            cell.alignment = Alignment(wrap_text=True)

    ws.freeze_panes = "A2"
    file_name = "mumbai_builders.xlsx"
    wb.save(file_name)
    print(f"\nSaved {len(records)} records -> {file_name}")

    # Fill rate report
    print(f"\n{'Column':<14} Filled / Total")
    print("-" * 35)
    for h in headers:
        filled = sum(1 for rec in records if rec.get(h))
        pct    = round(filled / len(records) * 100) if records else 0
        print(f"  {h:<12}  {filled:>5} / {len(records)}  ({pct}%)")


def main():
    records = scrape_builders(TARGET_URL)
    if not records:
        print("No data found.")
        return
    save_to_excel(records)


if __name__ == "__main__":
    main()